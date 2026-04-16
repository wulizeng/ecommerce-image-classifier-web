from typing import List, Dict, Any
import openpyxl


def read_links(file_path: str) -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        style_no, spu, skuid, link = (row[i] if i < len(row) else None for i in range(4))
        if link:
            rows.append({
                '款号': str(style_no or ''),
                'SPU': str(spu or ''),
                'SKUID': str(skuid or ''),
                '链接': str(link)
            })
    return rows


def write_results(results: List[Dict[str, Any]], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['款号', 'SPU', 'SKUID', '链接', '识别结果', '处理状态'])
    for item in results:
        ws.append([
            item.get('款号', ''),
            item.get('SPU', ''),
            item.get('SKUID', ''),
            item.get('链接', ''),
            item.get('label', ''),
            item.get('status', ''),
        ])
    wb.save(output_path)
